from pprint import pprint
from openpyxl import load_workbook

class HandleExcel:

    def __init__(self, file_name,sheet_name):
        #创建表对象
        self.wb = load_workbook(filename=file_name)
        #获取表单对象
        self.sheet_obj = self.wb[sheet_name]

    def get_cases(self):
        case_list = []
        cases = list(self.sheet_obj.iter_rows(values_only=True))
        title = cases[0]
        values = cases[1:]
        for case in values:
            case_dict = dict(list(zip(title,case)))
            case_list.append(case_dict)
        self.close_file()
        return case_list

    def close_file(self):
        self.wb.close()