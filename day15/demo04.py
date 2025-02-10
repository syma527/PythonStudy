"""
============================
Project: py52
Author:柠檬班-海励
Time:2022/7/20 21:17
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
"""
一、写excel文件【了解】
注意：写excel文件的时候要关闭excel，否则会报错
1、写单元格(覆盖)
sheet_obj["A1"] = "py52"
sheet_obj.cell(row=1,column=2,value="test_value")

2、追加写单元格(不会覆盖原有数据)
在excel最后行加入数据
可迭代对象的每个元素会占一个单元格，一个可迭代对象的数据占一行，不会换行
接收可迭代对象：list，tuple，dict，生成器
sheet_obj.append([1,2,3,4,5])
test_dict = {"A":"test1","B":"test2"}
"""
from openpyxl import load_workbook

wb = load_workbook(filename="cases.xlsx")
sheet_obj = wb["Sheet1"]

#写入数据(覆盖)
# sheet_obj["A1"] = "py52"
# sheet_obj.cell(row=1,column=2,value="test_value")

#追加写入(不覆盖数据)
#接收可迭代对象：list，tuple，dict，生成器
test_list = [1,2,3,4,5]
test_dict = {"A":"test1","B":"test2"}
sheet_obj.append(test_dict)

wb.save("cases.xlsx")
# print(sheet_obj["B1"].value)

wb.close()





