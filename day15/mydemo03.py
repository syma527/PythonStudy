from openpyxl import load_workbook


def read_excel_data(file_path):
    try:
        wb = load_workbook(filename=file_path)
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return []
    except Exception as e:
        print(f"Error: Failed to load workbook. {e}")
        return []

    if "Sheet1" not in wb.sheetnames:
        print("Error: Sheet 'Sheet1' not found in the workbook.")
        return []

    sheet_obj = wb["Sheet1"]

    result = list(sheet_obj.iter_rows(min_row=1, min_col=1, values_only=True))

    if not result:
        print("Error: The sheet is empty.")
        return []

    keep_list = []
    key_list = result[0]
    value_list = result[1:]

    for case in value_list:
        if len(key_list) != len(case):
            print("Error: Key and value lists do not match in length.")
            continue
        keep_list.append(dict(list(zip(key_list, case))))

    wb.close()

    return keep_list


file_path = "cases.xlsx"
keep_list = read_excel_data(file_path)
print(keep_list)
