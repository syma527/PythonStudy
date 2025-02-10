from openpyxl import load_workbook

wb = load_workbook(filename="cases.xlsx")
sheet_obj = wb["Sheet1"]
"""
result = sheet_obj.iter_cols(min_row=None,max_row=None,min_col=None, max_col=None,values_only=False)

for i in result:
    for j in i:
        print(j.value)

"""

from openpyxl import load_workbook

wb = load_workbook(filename="cases.xlsx")
if sheet_obj is None:
    raise ValueError("sheet_obj 未初始化")
sheet_obj = wb["Sheet1"]

cell_obj = sheet_obj.cell(row=2,column=4)
print(cell_obj.value)

try:
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column
    # result = sheet_obj.iter_rows(min_row=1,max_row=None,min_col=1,max_col=None,values_only=True)
    # for i in result:
    #     for j in i:
    #         print(j)
    result = sheet_obj.iter_rows(min_row=1,max_row=None,min_col=1,max_col=None,values_only=True)
    print(list(result))
except AttributeError as e:
    print(f"Error:{e}")

except Exception as e:
    print(f"Unexpected error:{e}")


