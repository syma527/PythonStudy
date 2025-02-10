from openpyxl import load_workbook
from pprint import pprint

class HandleExcel:

    def __init__(self,file_name,sheet_name):
        self.wb = load_workbook(file_name)
        self.sheet_obj = self.wb[sheet_name]

    def get_excel_data(self):
        case_list = []
        result = list(self.sheet_obj.iter_rows(values_only=True))
        key_list = result[0]
        value_list = result[1:]
        if len(key_list) != len(value_list):
            print("Error length is not equal to the length of two list")
        for case in value_list:

            case_list.append(dict(list(zip(key_list,case))))
            self.close_file()
        return case_list
    def close_file(self):
        self.wb.close()

if __name__ == '__main__':
    cl = HandleExcel(file_name="cases.xlsx",sheet_name="Sheet1")
    result = cl.get_excel_data()
    pprint(result)