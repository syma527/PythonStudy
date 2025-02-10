"""
============================
Project: py52
Author:柠檬班-海励
Time:2022/7/20 20:11
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
"""
一、安装第三方库
pip install openpyxl==3.0.7

二、excel组成
1、excel文件 == 表格对象   一个表格对象有多个表单对象       小区大门
2、表单 == 表单对象        一个表单对象有多单元格对象       单元门
3、单元格 == 单元格对象     单元格对象是excel的最小组成部分  你加的门

三、操作步骤
from openpyxl import load_workbook
#第一步创建表格对象
wb = load_workbook(filename="cases.xlsx")
#第二步获取表单对象
sheet_obj = wb["Sheet1"]
#第三步获取单元格对象
cell_obj = sheet_obj["B2"]
cell_obj2 = sheet_obj.cell(3,2)
print(cell_obj)
#第四步获取对应单元格的值
print(cell_obj.value)
print(cell_obj2.value)
#第五步关闭excel连接
wb.close()

四、行操作
1、result = sheet_obj.iter_rows()
min_row=None： 最小的行索引值(索引从1开始，必须是int类型，默认是1)
max_row=None：最大的行索引值(索引从1开始，必须是int类型，默认是最大的行)
min_col=None,：最小的列索引值(索引从1开始，必须是int类型，默认是1)
max_col=None：最大的列索引值(索引从1开始，必须是int类型，默认是最大的列)
values_only=False: True获取到的是value值，False拿到的是对象需要通过vaule属性获取对应的值
注意点：切片是闭区间，包含起始和结束索引对应的行的数据。

2、获取最大行
sheet_obj.max_row

3、获取最大列
sheet_obj.max_column


五、列操作
result = sheet_obj.iter_cols(min_col=1, max_col=3,values_only=True)
wb = load_workbook(filename="cases.xlsx",read_only=True)
read_only=False: 可以直接获取列的值
read_only=True: 如果是True，列切片不可以直接读取值


"""

from openpyxl import load_workbook

wb = load_workbook(filename="cases.xlsx",read_only=False)
#read_only=False: 如果是True，列切片不可以直接读取值
sheet_obj = wb["Sheet1"]
#行切片
# False拿到的是对象需要通过vaule属性获取对应的值
# result = sheet_obj.iter_rows(min_row=1, max_row=3, min_col=1, max_col=6, values_only=False)
# for i in result:
#     for j in i:
#         print(j.value)

#  True获取到的是value值，
# result = sheet_obj.iter_rows(min_row=1, max_row=3, min_col=1, max_col=6, values_only=True)
# print(list(result))


#列切片
# result = sheet_obj.iter_cols(min_col=1, max_col=3,values_only=False)
# for i in result:
#     for j in i:
#         print(j.value)

#read_only=False(必须是False): 可以直接获取列的值
result = sheet_obj.iter_cols(min_col=1, max_col=3,values_only=True)
print(list(result))
wb.close()

result = sheet_obj.iter_rows(min_row=1, max_row=3,min_col=None,max_col=None,values_only=True)
for i in result:
    for j in i:
        print(j.value)
