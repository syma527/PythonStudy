from openpyxl import load_workbook
from pprint import pprint


class HandleExcel:
    def __init__(self, file_name, sheet_name):
        self.wb = load_workbook(filename=file_name)

        self.sheet_obj = self.wb[sheet_name]

    def get_cases(self):
        case_list = []
        cases = list(self.sheet_obj.iter_rows(values_only=True))
        title = cases[0]
        values = cases[1:]
        for case in values:
            case_dict = dict(list(zip(title, case)))
            case_list.append(case_dict)
        self.close_file()
        return case_list

    def close_file(self):
        self.wb.close()


if __name__ == '__main__':
    cl = HandleExcel(file_name="cases.xlsx", sheet_name="Sheet1")
    result = cl.get_cases()
    pprint(result)
