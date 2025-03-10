"""
============================
Project: py52
Author:柠檬班-海励
Time:2022/7/20 20:39
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""

"""
线性脚本

"""
from openpyxl import load_workbook
wb = load_workbook(filename="cases.xlsx",read_only=False)
sheet_obj = wb["Sheet1"]
# [{"用例id":"1","用例名称":"登陆成功"},{}]
result = list(sheet_obj.iter_rows(values_only=True))
print(result)
data_list = []
title_list = result[0]
value_list = result[1:]
for case in value_list:
    data_list.append(dict(list(zip(title_list, case))))
print(data_list)
wb.close()




