"""
============================
Project: py52
Author:柠檬班-海励
Time:2022/7/20 20:51
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
"""
excel 封装  掌握


"""
from pprint import pprint
from openpyxl import load_workbook
class HandleExcel:
    def __init__(self,file_name,sheet_name):
        #创建表对象
        self.wb = load_workbook(filename=file_name)
        #创建sheet对象
        self.sheet_obj = self.wb[sheet_name]

    def get_cases(self):
        case_list = []
        cases = list(self.sheet_obj.iter_rows(values_only=True))
        title = cases[0] #获取表头
        values = cases[1:] #获取表数据
        for case in values:
            case_dict = dict(list(zip(title,case))) #将表数据和表头进行打包，变成dict
            case_list.append(case_dict) #将每一个测试用例加到list当中，集中返回
        self.close_file()
        return case_list

    def close_file(self):
        self.wb.close()

if __name__ == '__main__':
    cl = HandleExcel(file_name="cases.xlsx",sheet_name="Sheet1")
    result = cl.get_cases()
    pprint(result)

