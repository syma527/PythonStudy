"""
一、文件操作（纯文本文件）
1、动态拼接路径
res = os.path.dirname(os.path.abspath(__file__))



import os

res = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join()
print(res)




from openpyxl import load_workbook

try:

    wb = load_workbook(filename="cases.xlsx")
except Exception as e:
    print("未找到文件，请检查路径是否正确")

else:
    sheet_obj = wb["Sheet1"]
cell_obj = sheet_obj["A3"]

cell_obj2 = sheet_obj.cell(row=1, column=2)
print(cell_obj.value)
print(cell_obj2.value)
wb.close()



from openpyxl import load_workbook

wb = load_workbook(filename="cases.xlsx")

sheet_obj = wb["Sheet1"]

cell_obj = sheet_obj["A3"]
cell_obj2 = sheet_obj.cell(1,2)

print(cell_obj.value)
print(cell_obj2.value)

wb.close()
"""

with open(file="a.txt",mode = "r+",encoding="utf-8") as file:
    file.writelines("changtaiyiyiyantixi")


file = open(file="a.txt",mode = "r",encoding = "utf-8")

