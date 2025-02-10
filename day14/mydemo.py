"""
res1 = os.path.abspath(__file__)
res2 = os.path.dirname(res1)
res = os.path.join(res2, "a_new.txt")
print(res)

file = open(file="a_new.txt",mode="r",encoding="utf-8")
print(file.read())
file.close()

with open(file="a.txt",mode = "r",encoding="utf-8") as file:
    result = file.read()
    print(result)


with open(file="a_new.txt",mode="r",encoding="utf-8") as file:
    result = file.read()
    print(result)

with open(file = "a_new.txt",mode = "r",encoding="utf-8") as file:
    result = file.read()
    print(result)

# print(__file__)

import os

with open("a.txt",mode="r",encoding="utf-8") as file:
    res = file.readlines()
    print(res)





with open(file="a.txt", mode="w", encoding="utf-8") as file:
    file.write("3939234,\n,sdfsdfsdfs\n,dsfs,\ndf,sdfsd,fs,df,sd,f,sdf,,g,dfg,,dfg,dfg,df,g")


with open(file="a.txt", mode="r", encoding="utf-8") as file:
    res = file.read()
    print(res)
"""

from openpyxl import load_workbook
# 创建表格对象
wb = load_workbook(filename="cases.xlsx")
# 获取表单对象
sheet_obj = wb["Sheet1"]
# 获取单元格对象
cell_obj = sheet_obj["B1"]
cell_obj2 = sheet_obj.cell(3,2)
print(cell_obj.value)
print(cell_obj2.value)


