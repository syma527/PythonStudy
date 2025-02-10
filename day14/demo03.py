from openpyxl import load_workbook
"""
============================
Project: py52
Author:柠檬班-海励
Time:2022/7/18 21:32
E-mail:2227092769@qq.com
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
============================
"""
"""
excel操作使用openpyxl
作用：excel存测试数据

一、安装第三方库
pip install openpyxl==3.0.7

二、excel组成
1、excel文件 == 表格对象   一个表格对象有多个表单对象       小区大门
2、表单 == 表单对象        一个表单对象有多单元格对象       单元门
3、单元格 == 单元格对象     单元格对象是excel的最小组成部分  你加的门

三、操作步骤
from openpyxl import load_workbook
#第一步创建表格对象
wb = load_workbook(filename="cases.xlsx")
#第二步获取表单对象
sheet_obj = wb["Sheet1"]
#第三步获取单元格对象
cell_obj = sheet_obj["B2"]
cell_obj2 = sheet_obj.cell(3,2)
print(cell_obj)
#第四步获取对应单元格的值
print(cell_obj.value)
print(cell_obj2.value)
#第五步关闭excel连接
wb.close()

"""



wb = load_workbook(filename="cases.xlsx")


# 获取所有的表单名称
print(wb.sheetnames)
# 根据名称获取对应的表单对象
for i in wb.sheetnames:
    print(wb[i])

wb.close()
